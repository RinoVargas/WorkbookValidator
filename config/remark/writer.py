from config.finder import ValidatorFinder
from random import randrange
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

class RemarkWriter:
    
    __finder = None
    __workbook = None
    __worksheet = None
    __configuration = None
    
    def __init__(self, configuration):
        self.__configuration = configuration
        self.__finder = ValidatorFinder()
        self.__workbook = self.create_workbook(self.__configuration.SRC)
        self.__worksheet = self.__workbook.active
        
    def create_workbook(self, source):
        """
            Crear un Objeto 'Workbook' a partir del Objeto 'Dataframe'
            que poseen los datos
        """
        wb = Workbook()
        ws = wb.active

        for row in dataframe_to_rows(source):
            ws.append(row)

        # Eliminar columna de 'labels' creada por el Dataframe
        ws.delete_cols(1)
        # Eliminar fila en blanco que se inserta para separar la cabecera del cuerpo del Worksheet
        ws.delete_rows(2)
        
        return wb
        
    def get_column_identifier(self, column_name):
        """
            Retorna un diccionario donde las claves son los nombre de las columnas y los 
            valores son los números identificadores de estas en el Worksheet.
        """
        header_identifier = {cell.value : cell.column - 1 for cell in self.__worksheet[1]}
        
        return header_identifier[column_name]
        
    def list_writers(self):
        writers = dir(RemarkWriter)
        writers = tuple(filter(lambda x: x.startswith("write_from_"), writers))
        
        return writers
    
    def writer_wrapper(self, method_name):
        classname = self.__finder.get_classname_by_methodname(method_name)
        writers = self.list_writers()
        
        for writer in writers:
            
            if classname == writer.lstrip('write_from_'):
                
                return getattr(RemarkWriter ,writer)
            
        raise ValueError(f"'{writer}' writer methods does not exist.")
    
    def random_hexcolor(self):
        r = randrange(160,256)
        g = randrange(160,256)
        b = randrange(160,256)
        hex_rgb = tuple([r,g,b])
        
        return '%02x%02x%02x' % hex_rgb
    
    def insert_commet(self, row, message, validation_result):
        parameters = validation_result.copy()
        parameters.pop('records')
        message = message + "\n"
        message = message.format(**parameters)
        if row.comment is None:
            row.comment = Comment("","RecordValidator")
            
        comment_text = row.comment.text + message
        row.comment = Comment(comment_text, "RecordValidator")
        
        return row
        
    def write_worksheet(self, write_config, validation_result):
        columns_reference = write_config['columns_reference']
        writer_func = self.writer_wrapper(write_config['method_name'])
        hex_color = self.random_hexcolor()
        
        for column in columns_reference:
            writer_func(self, column, write_config, validation_result, hex_color)
    
    def export_to_excel(self):
        now = datetime.now()
        str_date = now.strftime("%Y-%m-%d_%H-%M-%S")
        filename = self.__configuration.ENV['source_file'].rstrip(".xls.xlsx")
        self.__workbook.save(self.__configuration.ENV['output_dir'] + "/" + filename + f"-EVALUATED-{str_date}" + ".xls")
    
    # Writers
    
    def write_from_CellValidator(self, column, write_config, validation_result, hex_color):
        id_column = self.__configuration.ENV['identification_column']
        id_column_num = self.get_column_identifier(id_column)
        column_num = self.get_column_identifier(column)
        validation_result['value'] = None
        validation_result['record'] = None
        
        for row in self.__worksheet.iter_rows(min_row=2):
                
            if row[id_column_num].value in validation_result['records']:
                validation_result['value'] = row[column_num].value
                validation_result['record'] = row[id_column_num].value
                
                # Insertar comentario en la celda
                if write_config['show'] is True:
                    self.insert_commet(row[column_num], write_config['message'], validation_result)
                    
                row[column_num].fill = PatternFill(start_color = hex_color, end_color = hex_color,
                                                   fill_type = "solid")
        
    def write_from_DuplicateValidator(self, column, write_config, validation_result, hex_color):
        id_column = self.__configuration.ENV['identification_column']
        id_column_num = self.get_column_identifier(id_column)
        column_num = self.get_column_identifier(column)
        method_type = write_config['configuration']['method_type']
        validation_result['value'] = None
        validation_result['record'] = None
        lock = False     
        
        # Recorrer Observaciones
        for record in validation_result['records']:
            color_group = self.random_hexcolor() if method_type == 'single_column' else hex_color
            
            # Recorrer Worksheet
            for row in self.__worksheet.iter_rows(min_row = 2):
                cell_value = row[id_column_num].value
                
                if method_type == 'single_column':
                    if cell_value in validation_result['records']:
                        validation_result['value'] = row[column_num].value
                        validation_result['record'] = row[id_column_num].value
                        lock = True
                        
                if method_type == 'double_column':
                    if cell_value == record['record']:
                        validation_result['value'] = row[column_num].value
                        validation_result['record'] = row[id_column_num].value
                        lock = True
                
                if lock is True:
                    
                    # Insertar comentario en la celda
                    if write_config['show'] is True:
                        self.insert_commet(row[column_num], write_config['message'], validation_result)

                    # Colorear el fondo de la celda
                    row[column_num].fill = PatternFill(start_color = color_group, 
                                                       end_color = color_group, 
                                                       fill_type = "solid")  
                    lock = False
                    
                    if method_type == 'double_column':
                        break